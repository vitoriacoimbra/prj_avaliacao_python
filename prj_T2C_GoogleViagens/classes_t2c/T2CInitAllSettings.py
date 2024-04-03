from openpyxl import load_workbook, Workbook
from pathlib import Path
from openpyxl.worksheet.worksheet import Worksheet

ROOT_DIR = Path(__file__).parent.parent

class T2CInitAllSettings:
    """
    Classe para carregar todas as configurações necessárias.

    Parâmetros:
    
    Retorna:
    """
    
    def load_config(self) -> dict:
        """
        Carrega o arquivo de configuração Config.xlsx e retorna um dicionário com as configurações.

        Parâmetros:

        Retorna:
        - dict: dicionário com as configurações lidas do arquivo Config.xlsx.
        """

        var_wbkConfig:Workbook = load_workbook(filename=ROOT_DIR.__str__() + "\\resources\\config\\Config.xlsx")
        var_wshtSettings:Worksheet = var_wbkConfig.get_sheet_by_name("Settings")
        var_wshtConstants:Worksheet = var_wbkConfig.get_sheet_by_name("Constants")
        var_wshtCredentials:Worksheet = var_wbkConfig.get_sheet_by_name("Credentials")

        #Iniciando o dicionário
        arg_dictConfig = dict()

        #Loop adicionando dados da aba settings para o dicionário
        var_intTotalRows = var_wshtSettings.max_row
        for var_intRowNumber in range(1, var_intTotalRows):
            var_gncDictKey = var_wshtSettings["A" + (var_intRowNumber+1).__str__()].value
            var_gncDictObj = var_wshtSettings["B" + (var_intRowNumber+1).__str__()].value
            #Apenas inclui no dicionário se as duas colunas não forem nulas
            if(var_gncDictKey is not None and var_gncDictObj is not None): arg_dictConfig[var_gncDictKey] = var_gncDictObj 

        #Mesmo loop, agora com constants
        var_intTotalRows = var_wshtConstants.max_row
        for var_intRowNumber in range(1, var_intTotalRows):
            var_gncDictKey = var_wshtConstants["A" + (var_intRowNumber+1).__str__()].value
            var_gncDictObj = var_wshtConstants["B" + (var_intRowNumber+1).__str__()].value
            #Apenas inclui no dicionário se as duas colunas não forem nulas
            if(var_gncDictKey is not None and var_gncDictObj is not None): arg_dictConfig[var_gncDictKey] = var_gncDictObj 

        #Mesmo loop, agora com credentials
        var_intTotalRows = var_wshtCredentials.max_row
        for var_intRowNumber in range(1, var_intTotalRows):
            var_gncDictKey = var_wshtCredentials["A" + (var_intRowNumber+1).__str__()].value
            var_gncDictObj = var_wshtCredentials["B" + (var_intRowNumber+1).__str__()].value
            #Apenas inclui no dicionário se as duas colunas não forem nulas
            if(var_gncDictKey is not None and var_gncDictObj is not None): arg_dictConfig[var_gncDictKey] = var_gncDictObj 

        #Retorna o objeto dicionario para ser usado no código
        return arg_dictConfig 
