import shutil, os
from pathlib import Path
from openpyxl import load_workbook

ROOT_DIR = Path(__file__).parent.parent.parent.__str__()
PATH_TEMPLATE_ANALITICO = ROOT_DIR + "\\resources\\templates\\Relatorio_Analitico.xlsx"
PATH_TEMPLATE_SINTETICO = ROOT_DIR + "\\resources\\templates\\Relatorio_Sintetico.xlsx"

class T2CRelatorios:
    """
    Classe responsável por manipular relatórios, incluindo linhas e preenchendo a partir de templates
    
    Parâmetros:

    Retorna:

    """
    
    def __init__(self, arg_dictConfig:dict):
        """
        Inicializa a classe T2CRelatorios.
        Copia os templates dos relatórios e deixa eles prontos na pasta de saída, possível de acessar através de propriedades da classe também

        Parâmetros:
        - arg_dictConfig (dict): dicionário de configuração.

        Retorna:

        """

        #Fazendo uma cópia do relatório template e deixando na pasta de saída
        self.var_strPathRelatorioAnalitico = arg_dictConfig["CaminhoPastaRelatorios"] + "Relatorio_Analitico_" + arg_dictConfig["NomeProcesso"] + ".xlsx"
        self.var_strPathRelatorioSintetico = arg_dictConfig["CaminhoPastaRelatorios"] + "Relatorio_Sintetico_" + arg_dictConfig["NomeProcesso"] + ".xlsx"

        #Copiando apenas se não existir
        if(not os.path.exists(self.var_strPathRelatorioAnalitico)):
            shutil.copy(src=PATH_TEMPLATE_ANALITICO, dst=self.var_strPathRelatorioAnalitico)
        
            #Colocando nome do processo no relatório
            var_wbkAnalitico = load_workbook(self.var_strPathRelatorioAnalitico)
            var_wshtAnalitico = var_wbkAnalitico.active
            var_wshtAnalitico["D4"] = arg_dictConfig["NomeProcesso"]

            var_wbkAnalitico.save(self.var_strPathRelatorioAnalitico)
            var_wbkAnalitico.close()

        if(not os.path.exists(self.var_strPathRelatorioSintetico)):
            shutil.copy(src=PATH_TEMPLATE_SINTETICO, dst=self.var_strPathRelatorioSintetico)
        
            #Colocando nome do processo no relatório
            var_wbkSintetico = load_workbook(self.var_strPathRelatorioSintetico)
            var_wshtSintetico = var_wbkSintetico.active
            var_wshtSintetico["C4"] = arg_dictConfig["NomeProcesso"]

            var_wbkSintetico.save(self.var_strPathRelatorioSintetico)
            var_wbkSintetico.close()

    def inserir_linha_analitico(self, arg_listValores:list):
        """
        Insere uma linha nova no relatório analítico com os valores fornecidos.

        Parâmetros:
        - arg_listValores (list): lista de valores.

        Retorna:

        """
        var_wbkAnalitico = load_workbook(self.var_strPathRelatorioAnalitico)
        var_wshtAnalitico = var_wbkAnalitico.active

        var_intIndexNewline:int = None
        var_intIndexAux = 5
        #Encontrando linha vazia
        while(var_intIndexNewline is None):
            if(var_wshtAnalitico["A" + var_intIndexAux.__str__()].value is None):
                var_intIndexNewline = var_intIndexAux
            else:
                var_intIndexAux += 1
        #Escrevendo 
        for var_intColumn in range(len(arg_listValores)):
            var_wshtAnalitico.cell(row=var_intIndexNewline, column=var_intColumn+1, value=arg_listValores[var_intColumn])

        var_wbkAnalitico.save(self.var_strPathRelatorioAnalitico)
        var_wbkAnalitico.close()

    def inserir_linha_sintetico(self, arg_listValores:list):
        """
        Insere uma linha nova no relatório sintético com os valores fornecidos.

        Parâmetros:
        - arg_listValores (list): lista de valores.

        Retorna:
        
        """
        var_wbkSintetico = load_workbook(self.var_strPathRelatorioSintetico)
        var_wshtSintetico = var_wbkSintetico.active

        var_intIndexNewline:int = None
        var_intIndexAux = 5
        #Encontrando linha vazia
        while(var_intIndexNewline is None):
            if(var_wshtSintetico["A" + var_intIndexAux.__str__()].value is None):
                var_intIndexNewline = var_intIndexAux
            else:
                var_intIndexAux += 1
        #Escrevendo 
        for var_intColumn in range(len(arg_listValores)):
            var_wshtSintetico.cell(row=var_intIndexNewline, column=var_intColumn+1, value=arg_listValores[var_intColumn])

        var_wbkSintetico.save(self.var_strPathRelatorioSintetico)
        var_wbkSintetico.close()

